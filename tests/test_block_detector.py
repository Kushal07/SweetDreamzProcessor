from openpyxl import Workbook

from processor.block_detector import (
    BlockDetector,
    BlockState,
)


def test_empty_block():

    wb = Workbook()
    ws = wb.active

    detector = BlockDetector()

    state = detector.detect(
        worksheet=ws,
        row=2,
        start_column=4,
    )

    assert state is BlockState.EMPTY


def test_complete_block():

    wb = Workbook()
    ws = wb.active

    ws.cell(row=2, column=4).value = "4,7,9"
    ws.cell(row=2, column=5).value = "52,52,52"
    ws.cell(row=2, column=6).value = "31,65,88"

    detector = BlockDetector()

    state = detector.detect(
        worksheet=ws,
        row=2,
        start_column=4,
    )

    assert state is BlockState.COMPLETE


def test_partial_block():

    wb = Workbook()
    ws = wb.active

    ws.cell(row=2, column=4).value = "4,7,9"
    ws.cell(row=2, column=5).value = "52,52,52"

    detector = BlockDetector()

    state = detector.detect(
        worksheet=ws,
        row=2,
        start_column=4,
    )

    assert state is BlockState.PARTIAL