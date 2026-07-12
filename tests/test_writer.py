from openpyxl import Workbook

from processor.writer import WorkbookWriter


def test_write_middle_block():
    writer = WorkbookWriter()

    wb = Workbook()
    ws = wb.active

    writer.write_middle_block(
        worksheet=ws,
        row=2,
        start_column=4,
        data={
            "series": "4,7,9",
            "pair": "52,52,52",
            "blank": "31,65,88",
        },
    )

    assert ws["D2"].value == "4,7,9"
    assert ws["E2"].value == "52,52,52"
    assert ws["F2"].value == "31,65,88"