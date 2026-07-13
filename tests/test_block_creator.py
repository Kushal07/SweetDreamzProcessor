from openpyxl import Workbook

from processor.block_creator import BlockCreator

def test_fill_middle_header():

    workbook = Workbook()

    worksheet = workbook.active

    creator = BlockCreator()

    creator.fill_middle_header(
        worksheet=worksheet,
        start_column=4,
        pair="52",
    )

    assert worksheet.cell(1, 4).value == "Series"
    assert worksheet.cell(1, 5).value == "52"
    assert worksheet.cell(1, 6).value == "Blank"

def test_fill_last_header():

    workbook = Workbook()

    worksheet = workbook.active

    creator = BlockCreator()

    creator.fill_last_header(
        worksheet=worksheet,
        start_column=7,
        pair="18",
    )

    assert worksheet.cell(1, 7).value == "Series"
    assert worksheet.cell(1, 8).value == "Blank"
    assert worksheet.cell(1, 9).value == "18"

