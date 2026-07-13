from openpyxl import Workbook

from processor.workbook_mapper import WorkbookMapper


def test_map_pair_columns_with_two_valid_blocks():

    wb = Workbook()

    ws = wb.active

    ws["D1"] = "Series"
    ws["E1"] = "00"
    ws["F1"] = "Last"

    ws["G1"] = "Series"
    ws["H1"] = "01"
    ws["I1"] = "Last"

    mapper = WorkbookMapper()

    mapping = mapper.map_pair_columns(ws)

    assert mapping["00"]["series"] == 4
    assert mapping["00"]["pair"] == 5
    assert mapping["00"]["last"] == 6

    assert mapping["01"]["series"] == 7
    assert mapping["01"]["pair"] == 8
    assert mapping["01"]["last"] == 9

    assert len(mapping) == 2

    assert "00" in mapping
    assert "01" in mapping

def test_get_missing_pairs():

    mapper = WorkbookMapper()

    mapping = {
        "00": {},
        "01": {},
        "03": {},
    }

    missing = mapper.get_missing_pairs(mapping)

    assert "02" in missing
    assert "00" not in missing
    assert "01" not in missing
    assert "03" not in missing
    assert len(missing) == 97


def test_find_reserved_space_ignores_source_columns():

    workbook = Workbook()
    worksheet = workbook.active

    # Source worksheet headers
    worksheet["A1"] = "Date"
    worksheet["B1"] = "1st Prize"
    worksheet["C1"] = "2nd Prize"

    # Blank cells before arrangement
    worksheet["D1"] = ""
    worksheet["E1"] = ""
    worksheet["F1"] = ""

    # Arrangement starts here
    worksheet["G1"] = "Series"
    worksheet["H1"] = "00"
    worksheet["I1"] = "Blank"

    mapper = WorkbookMapper()

    reserved = mapper.find_reserved_space(
        worksheet=worksheet,
        pair="02",
    )
    assert reserved is None

def test_find_insert_position_before_existing_pair():

    mapper = WorkbookMapper()

    mapping = {
        "00": {"series": 4, "pair": 5, "last": 6},
        "01": {"series": 7, "pair": 8, "last": 9},
        "03": {"series": 13, "pair": 14, "last": 15},
    }

    column = mapper.find_insert_position(
        mapping=mapping,
        pair="02",
    )

    assert column == 13


def test_find_insert_position_append_to_end():

    mapper = WorkbookMapper()

    mapping = {
        "98": {"series": 298, "pair": 299, "last": 300},
    }

    column = mapper.find_insert_position(
        mapping=mapping,
        pair="99",
    )

    assert column == 301

import pytest


def test_find_insert_position_empty_mapping():

    mapper = WorkbookMapper()

    with pytest.raises(ValueError):
        mapper.find_insert_position(
            mapping={},
            pair="00",
        )