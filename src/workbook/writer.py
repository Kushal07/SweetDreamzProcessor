from openpyxl.styles import PatternFill, Alignment
class WorkbookWriter:
    def __init__(self):
        self.fill_s = PatternFill(start_color='99D6EA', end_color='99D6EA', fill_type='solid')
        self.fill_y = PatternFill(start_color='FFB600', end_color='FFB600', fill_type='solid')
        self.fill_g = PatternFill(start_color='46AB21', end_color='46AB21', fill_type='solid')
        self.align = Alignment(wrap_text=True, vertical='top', horizontal='center')
    def write_block(self, ws, r, c, data, is_ld=False):
        c1 = ws.cell(r, c, value=data[0]); c1.fill, c1.alignment = self.fill_s, self.align
        c2, c3 = ws.cell(r, c+1, value=data[1]), ws.cell(r, c+2, value=data[2])
        if not is_ld: c2.fill, c3.fill = self.fill_y, self.fill_g
        else: c2.fill, c3.fill = self.fill_g, self.fill_y
        c2.alignment, c3.alignment = self.align, self.align
