from openpyxl.styles import PatternFill, Font, Alignment
class GridManager:
    def __init__(self, wb):
        self.wb = wb
        self.fill_s = PatternFill(start_color='99D6EA', end_color='99D6EA', fill_type='solid')
        self.fill_y = PatternFill(start_color='FFB600', end_color='FFB600', fill_type='solid')
        self.fill_g = PatternFill(start_color='46AB21', end_color='46AB21', fill_type='solid')
        self.h_font = Font(bold=True, size=10)
        self.align = Alignment(wrap_text=True, vertical='top', horizontal='center')

    def recreate_clean_sheet(self, old_name):
        old_ws = self.wb[old_name]
        new_ws = self.wb.create_sheet(old_name + '_CLEAN')
        new_ws.freeze_panes = 'B2'
        new_ws.column_dimensions['A'].width, new_ws.column_dimensions['B'].width, new_ws.column_dimensions['C'].width = 30, 45, 45
        
        for c, txt in enumerate(['Date', '1st Prize', '2nd Prize'], 1):
            cell = new_ws.cell(1, c, value=txt)
            cell.font, cell.alignment = self.h_font, self.align
        
        is_ld = 'last' in old_name.lower() or 'digit' in old_name.lower()
        for i in range(100):
            pv, base = str(i).zfill(2), 4 + (i * 3)
            if not is_ld: h_data = [(f'S_{pv}', self.fill_s), (pv, self.fill_y), (f'B_{pv}', self.fill_g)]
            else: h_data = [(f'S_{pv}', self.fill_s), (f'B_{pv}', self.fill_g), (pv, self.fill_y)]
            for idx, (txt, fill) in enumerate(h_data):
                c = new_ws.cell(1, base + idx, value=txt)
                c.font, c.alignment, c.fill = self.h_font, self.align, fill
                new_ws.column_dimensions[c.column_letter].width = 10
        del self.wb[old_name]; new_ws.title = old_name; return new_ws
