from openpyxl import load_workbook
from ..business.models import LotteryNumber
from ..business.processor import NumberExtractor
from ..workbook.grid_manager import GridManager
from ..workbook.writer import WorkbookWriter

class SweetDreamzController:
    def __init__(self): self.wb = None
    def process(self, path, callback):
        self.wb = load_workbook(path)
        ex, wr, grid = NumberExtractor(), WorkbookWriter(), GridManager(self.wb)
        names = list(self.wb.sheetnames)
        for s_idx, sn in enumerate(names):
            old_ws = self.wb[sn]; src_data = []
            for r in range(2, old_ws.max_row + 1):
                src_data.append((old_ws.cell(r,1).value, old_ws.cell(r,2).value, old_ws.cell(r,3).value))
            is_ld = 'last' in sn.lower() or 'digit' in sn.lower()
            ws = grid.recreate_clean_sheet(sn)
            for r_idx, (d_v, p1_v, p2_v) in enumerate(src_data):
                rn = r_idx + 2
                ws.cell(rn, 1, value=d_v).number_format = 'dddd, dd mmmm yyyy'
                ws.cell(rn, 1).alignment = grid.align
                ws.cell(rn, 2, value=ex.format_prize_text(p1_v, 2)).alignment = grid.align
                ws.cell(rn, 3, value=ex.format_prize_text(p2_v, 3)).alignment = grid.align
                p1_n, p2_n = [LotteryNumber.from_str(n) for n in ex.extract_numbers(p1_v)], [LotteryNumber.from_str(n) for n in ex.extract_numbers(p2_v)]
                if not p1_n and not p2_n: continue
                row_map = {}
                for p in (p1_n + p2_n):
                    key = p.last_pair if is_ld else p.middle_pair
                    if key not in row_map: row_map[key] = ([], [], [])
                    row_map[key][0].append(p.series); row_map[key][1].append(p.middle_pair); row_map[key][2].append(p.last_pair)
                for k, v in row_map.items():
                    wr.write_block(ws, rn, 4+(int(k)*3), (','.join(v[0]), ','.join(v[1]), ','.join(v[2])), is_ld)
                if rn % 10 == 0: callback((s_idx/len(names)) + ((r_idx/len(src_data))/len(names))*100, f"Processing {sn}...")
        callback(100, "Done!")
        return True
    def save(self, path): self.wb.save(path)
