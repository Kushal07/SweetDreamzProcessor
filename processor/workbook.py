"""
Workbook Manager
================

Handles loading and validating Excel workbooks.
"""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from utils.logger import get_logger
from processor.validator import WorkbookValidator
from openpyxl.worksheet.worksheet import Worksheet

class WorkbookManager:
    """Handles workbook loading."""

    def __init__(self) -> None:
        self.logger = get_logger(__name__)
        self.workbook: Workbook | None = None
        self.file_path: Path | None = None
        self.validator = WorkbookValidator()

    def load(self, filename: str | Path) -> Workbook:
        """
        Load workbook.

        Raises
        ------
        FileNotFoundError
        ValueError
        """

        self.file_path = Path(filename)

        if not self.file_path.exists():
            raise FileNotFoundError(self.file_path)

        self.logger.info(f"Loading workbook: {self.file_path.name}")

        self.workbook = load_workbook(
            filename=self.file_path,
            data_only=False
        )
        # Validate AFTER the workbook is loaded
        self.validator.validate(self.workbook)

        self.logger.info("Workbook loaded successfully.")

        return self.workbook

    def sheet_names(self) -> List[str]:
        """Return worksheet names."""

        if self.workbook is None:
            return []

        return self.workbook.sheetnames

    def active_sheet(self):
        """Return active worksheet."""

        if self.workbook is None:
            raise RuntimeError("Workbook not loaded.")

        return self.workbook.active
    
    def get_sheet(self, sheet_name: str) -> Worksheet:
        """Return a worksheet by name."""

        if self.workbook is None:
            raise RuntimeError("Workbook not loaded.")

        return self.workbook[sheet_name]

    def information(self) -> dict:
        """Return workbook information."""

        if self.workbook is None:
            raise RuntimeError("Workbook not loaded.")

        ws = self.active_sheet()

        return {
            "filename": self.file_path.name,
            "worksheets": len(self.workbook.sheetnames),
            "sheet_names": self.workbook.sheetnames,
            "rows": ws.max_row,
            "columns": ws.max_column,
        }
    
    def suggested_output_filename(self) -> Path:
        """
        Return the suggested filename for the processed workbook.

        Example
        -------
        Lottery.xlsx

        becomes

        Lottery_Modified_KB.xlsx
        """

        if self.file_path is None:
            raise RuntimeError("Workbook path not available.")

        return self.file_path.with_name(
            f"{self.file_path.stem}_Modified_KB{self.file_path.suffix}"
        )

    def get_row_data(
        self,
        sheet_name: str,
        row_number: int,
        date_col: int = 1,
        first_prize_col: int = 2,
        second_prize_col: int = 3,
    ) -> dict:
        """
        Read one row from a worksheet.

        Returns
        -------
        dict
        """

        ws = self.get_sheet(sheet_name)

        return {
            "date": ws.cell(row=row_number, column=date_col).value,
            "first_prize": ws.cell(row=row_number, column=first_prize_col).value,
            "second_prize": ws.cell(row=row_number, column=second_prize_col).value,
        } 
    
    def save(self, filename: str | Path) -> None:
        """
        Save the currently loaded workbook.
        """

        if self.workbook is None:
            raise RuntimeError("Workbook not loaded.")

        filename = Path(filename)

        filename.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        self.workbook.save(filename)

        self.logger.info(f"Workbook saved: {filename}")