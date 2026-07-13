"""
SweetDreamzProcessor
Workbook Mapper
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet


class WorkbookMapper:
    """
    Builds a mapping of Pair -> Worksheet Columns.

    Example:

    {
        "52": {
            "series": 154,
            "pair": 155,
            "last": 156,
        }
    }
    """

    def is_complete_mapping(
        self,
        mapping: dict[str, dict[str, int]],
    ) -> bool:
        """
        Return True only if all arrangement
        pairs 00-99 exist.
        """

        return len(mapping) == 100
    
    def get_missing_pairs(
        self,
        mapping: dict[str, dict[str, int]],
    ) -> list[str]:
        """
        Return all missing arrangement pairs.

        Example

            Existing:
                00
                01
                03

            Returns:
                ["02", "04", ..., "99"]
        """

        missing: list[str] = []

        for pair in range(100):

            pair_key = f"{pair:02d}"

            if pair_key not in mapping:
                missing.append(pair_key)

        return missing


    def map_pair_columns(
        self,
        worksheet: Worksheet,
        header_row: int = 1,
    ) -> dict[str, dict[str, int]]:

        mapping: dict[str, dict[str, int]] = {}

        max_col = worksheet.max_column

        column = 1

        while column <= max_col - 2:

            series = worksheet.cell(row=header_row, column=column).value
            pair = worksheet.cell(row=header_row, column=column + 1).value

            header = str(series).strip().lower()

            if header == "series" and pair is not None:

                # Convert Excel numeric values like 0.0 -> "00"
                try:
                    pair_key = f"{int(float(pair)):02d}"
                except (TypeError, ValueError):
                    pair_key = str(pair).strip().zfill(2)

                mapping[pair_key] = {
                    "series": column,
                    "pair": column + 1,
                    "last": column + 2,
                }

                column += 3

            else:
                column += 1

        return mapping
    
    def map_last_digit_columns(
        self,
        worksheet: Worksheet,
        header_row: int = 1,
    ) -> dict[str, dict[str, int]]:
        """
        Build a mapping for the Last Digit Arrangement worksheet.

        Expected header layout:

        S_00 | B_00 | 00
        S_01 | B_01 | 01
        ...
        """

        mapping: dict[str, dict[str, int]] = {}

        max_col = worksheet.max_column
        column = 1

        while column <= max_col - 2:

            header1 = str(
                worksheet.cell(
                    row=header_row,
                    column=column,
                ).value or ""
            ).strip()

            header2 = str(
                worksheet.cell(
                    row=header_row,
                    column=column + 1,
                ).value or ""
            ).strip()

            header3 = str(
                worksheet.cell(
                    row=header_row,
                    column=column + 2,
                ).value or ""
            ).strip()

            is_series = (
                header1.strip().lower() == "series"
                or header1.upper().startswith("S_")
            )

            is_blank = (
                header2.strip().lower() == "blank"
                or header2.upper().startswith("B_")
            )

            if is_series and is_blank:

                try:
                    pair_key = f"{int(float(header3)):02d}"
                except (TypeError, ValueError):
                    pair_key = header3.zfill(2)

                mapping[pair_key] = {
                    "series": column,
                    "blank": column + 1,
                    "pair": column + 2,
                }

                column += 3

            else:
                column += 1

        return mapping
    
    def find_arrangement_start(
        self,
        worksheet: Worksheet,
        header_row: int = 1,
    ) -> int | None:
        """
        Return the starting column of the first
        arrangement section.

        Returns None if no arrangement section exists.
        """

        max_col = worksheet.max_column

        for column in range(1, max_col + 1):

            value = worksheet.cell(
                row=header_row,
                column=column,
            ).value

            if str(value).strip().lower() == "series":
                return column

        return None
    
    def find_reserved_space(
        self,
        worksheet: Worksheet,
        pair: str,
        is_last_digit: bool = False,
        header_row: int = 1,
    ) -> int | None:
        """
        Find an existing empty 3-column header block that can be
        reused for the specified pair.

        Returns
        -------
        int | None
            Starting column of the reserved block, or None if
            no reserved space exists.
        """

        max_col = worksheet.max_column

        arrangement_start = self.find_arrangement_start(
            worksheet,
            header_row,
        )

        if arrangement_start is None:
            return None

        for column in range(
            arrangement_start,
            max_col - 1,
        ):

            first = worksheet.cell(
                row=header_row,
                column=column,
            ).value

            second = worksheet.cell(
                row=header_row,
                column=column + 1,
            ).value

            third = worksheet.cell(
                row=header_row,
                column=column + 2,
            ).value

            if (
                first in (None, "")
                and second in (None, "")
                and third in (None, "")
            ):
                return column

        return None
    

    def find_insert_position(
        self,
        mapping: dict[str, dict[str, int]],
        pair: str,
    ) -> int:
        """
        Return the worksheet column where a missing
        arrangement block should be inserted.

        The block is inserted before the first
        existing pair greater than the requested pair.

        If the pair belongs at the end, return the
        column immediately after the last block.
        """

        target = int(pair)

        if not mapping:
            raise ValueError(
                "Cannot determine insert position from an empty mapping."
    )

        existing_pairs = sorted(
            int(key)
            for key in mapping.keys()
        )

        for existing in existing_pairs:

            if existing > target:

                existing_key = f"{existing:02d}"

                return mapping[existing_key]["series"]

        last_pair = f"{existing_pairs[-1]:02d}"

        return mapping[last_pair]["last"] + 1