"""
SweetDreamzProcessor
Block Detector
"""

from __future__ import annotations

from enum import Enum

from openpyxl.worksheet.worksheet import Worksheet


class BlockState(Enum):
    """
    State of a 3-column arrangement block.
    """

    EMPTY = "empty"
    COMPLETE = "complete"
    PARTIAL = "partial"


class BlockDetector:
    """
    Detects the state of one 3-column block.

    Block layout:

    Series | Pair | Last
    """

    def detect(
        self,
        worksheet: Worksheet,
        row: int,
        start_column: int,
    ) -> BlockState:

        values = [
            worksheet.cell(row=row, column=start_column).value,
            worksheet.cell(row=row, column=start_column + 1).value,
            worksheet.cell(row=row, column=start_column + 2).value,
        ]

        filled = sum(
            value not in (None, "")
            for value in values
        )

        if filled == 0:
            return BlockState.EMPTY

        if filled == 3:
            return BlockState.COMPLETE

        return BlockState.PARTIAL