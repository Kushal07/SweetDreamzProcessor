"""
SweetDreamzProcessor
Block Creator
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet


class BlockCreator:
    """
    Creates or completes individual arrangement header blocks.

    Responsibilities
    ----------------
    - Create a new Number Wise header block.
    - Create a new Last Digit header block.
    - Complete an existing Number Wise header block.
    - Complete an existing Last Digit header block.

    This class does NOT:

    - Detect missing headers.
    - Decide where headers belong.
    - Refresh worksheet mappings.
    - Write lottery numbers.

    Those responsibilities belong to other components.
    """

    HEADER_ROW = 1

    # ------------------------------------------------------------------
    # Private Helpers
    # ------------------------------------------------------------------

    def _write_middle_header(
        self,
        worksheet: Worksheet,
        start_column: int,
        pair: str,
    ) -> None:
        """
        Write a Number Wise header block.

        Layout

            Series | Pair | Blank
        """

        worksheet.cell(
            row=self.HEADER_ROW,
            column=start_column,
            value="Series",
        )

        worksheet.cell(
            row=self.HEADER_ROW,
            column=start_column + 1,
            value=pair,
        )

        worksheet.cell(
            row=self.HEADER_ROW,
            column=start_column + 2,
            value="Blank",
        )

    def _write_last_header(
        self,
        worksheet: Worksheet,
        start_column: int,
        pair: str,
    ) -> None:
        """
        Write a Last Digit header block.

        Layout

            Series | Blank | Pair
        """

        worksheet.cell(
            row=self.HEADER_ROW,
            column=start_column,
            value="Series",
        )

        worksheet.cell(
            row=self.HEADER_ROW,
            column=start_column + 1,
            value="Blank",
        )

        worksheet.cell(
            row=self.HEADER_ROW,
            column=start_column + 2,
            value=pair,
        )

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def create_middle_header(
        self,
        worksheet: Worksheet,
        start_column: int,
        pair: str,
    ) -> None:
        """
        Create a new Number Wise header block.
        """

        self._write_middle_header(
            worksheet,
            start_column,
            pair,
        )

    def create_last_header(
        self,
        worksheet: Worksheet,
        start_column: int,
        pair: str,
    ) -> None:
        """
        Create a new Last Digit header block.
        """

        self._write_last_header(
            worksheet,
            start_column,
            pair,
        )

    def fill_middle_header(
        self,
        worksheet: Worksheet,
        start_column: int,
        pair: str,
    ) -> None:
        """
        Complete an existing Number Wise header block.
        """

        self._write_middle_header(
            worksheet,
            start_column,
            pair,
        )

    def fill_last_header(
        self,
        worksheet: Worksheet,
        start_column: int,
        pair: str,
    ) -> None:
        """
        Complete an existing Last Digit header block.
        """

        self._write_last_header(
            worksheet,
            start_column,
            pair,
        )

    # ------------------------------------------------------------------
    # Future ARR-004 Features
    # ------------------------------------------------------------------

    def insert_middle_block(
        self,
        worksheet: Worksheet,
        start_column: int,
        pair: str,
    ) -> None:
        """
        Insert a new Number Wise arrangement block.

        Not implemented yet.
        """
        worksheet.insert_cols(
            start_column,
            amount=3,
        )

        self.create_middle_header(
            worksheet=worksheet,
            start_column=start_column,
            pair=pair,
        )

    def insert_last_block(
        self,
        worksheet: Worksheet,
        start_column: int,
        pair: str,
    ) -> None:
        """
        Insert a new Last Digit arrangement block.

        Not implemented yet.
        """
        worksheet.insert_cols(
            start_column,
            amount=3,
        )

        self.create_last_header(
            worksheet=worksheet,
            start_column=start_column,
            pair=pair,
        )