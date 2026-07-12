"""
Workbook Validator
==================

Validates the structure of a Sweet Dreamz workbook.
"""

from __future__ import annotations

from openpyxl.workbook.workbook import Workbook


class WorkbookValidationError(Exception):
    """Raised when the workbook structure is invalid."""


class WorkbookValidator:
    """Validate workbook structure."""

    REQUIRED_SHEETS = (
        "Number wise arrangement",
        "Last digit arrangement",
    )

    REQUIRED_HEADERS = (
        "Date",
        "1st",
        "2nd",
    )

    def validate(self, workbook: Workbook) -> None:
        """
        Validate workbook.

        Raises
        ------
        WorkbookValidationError
        """

        # Required worksheets
        for sheet in self.REQUIRED_SHEETS:
            if sheet not in workbook.sheetnames:
                raise WorkbookValidationError(
                    f"Worksheet '{sheet}' not found."
                )

        ws = workbook["Number wise arrangement"]

        headers = [
            str(ws["A1"].value).strip(),
            str(ws["B1"].value).strip(),
            str(ws["C1"].value).strip(),
        ]

        if tuple(headers) != self.REQUIRED_HEADERS:
            raise WorkbookValidationError(
                f"Invalid headers: {headers}"
            )