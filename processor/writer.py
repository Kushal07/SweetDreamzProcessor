"""
Workbook Writer
"""

from openpyxl.worksheet.worksheet import Worksheet


class WorkbookWriter:
    """
    Writes arrangement data into worksheets.
    """

    def write_middle_block(
        self,
        worksheet: Worksheet,
        row: int,
        start_column: int,
        data: dict,
    ) -> None:
        """
        Write one 3-column middle pair block.

        start_column = first column of the block

        Example

        D = Series
        E = Pair
        F = Blank
        """

        worksheet.cell(
            row=row,
            column=start_column,
            value=data["series"],
        )

        worksheet.cell(
            row=row,
            column=start_column + 1,
            value=data["pair"],
        )

        worksheet.cell(
            row=row,
            column=start_column + 2,
            value=data["blank"],
        )

    def write_last_block(
        self,
        worksheet: Worksheet,
        row: int,
        start_column: int,
        data: dict,
    ) -> None:
        """
        Write one 3-column last digit block.

        start_column = first column of the block

        Example

        D = Series
        E = Blank (Middle Pair)
        F = Last Pair
        """

        worksheet.cell(
            row=row,
            column=start_column,
            value=data["series"],
        )

        worksheet.cell(
            row=row,
            column=start_column + 1,
            value=data["blank"],
        )

        worksheet.cell(
            row=row,
            column=start_column + 2,
            value=data["pair"],
        )